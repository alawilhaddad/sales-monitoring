from tkinter.filedialog import *
from tkinter.messagebox import showinfo
from openpyxl import Workbook, load_workbook
from excel.helper import convert_slash
from excel import var


class Excel:
    def __init__(self):
        self.path = None
        self.filename = None
        self.workbook = None
        self.sheet_list = []
        self.active_sheet = None
        self.state = False


class LoadExcel(Excel):
    def __init__(self):
        super().__init__()
        self.image = None
        self.options = None
        self.var = None
        self.cell = None
        self.month = []
        self.label = None
        self.selected_sheet = None

    def load(self, app, options, label):
        # Filedialog open
        self.path = askopenfilename(initialdir=var.default_open_dir,
                                    title='Open File',
                                    filetypes=(("Excel File", "*xlsx"),
                                               ("All Files", "*.*")))
        self.path = convert_slash(self.path)
        if self.path:
            self.workbook = load_workbook(self.path, data_only=True)
            self.load_attribute(app, options, label)

        else:
            return

    def load_attribute(self, app, options, label):
        self.filename = self.path.split('\\')[-1]
        self.sheet_list = self.workbook.sheetnames
        self.active_sheet = self.sheet_list[0]
        app.canvas.itemconfigure(label, text=self.filename)
        options['values'] = self.sheet_list
        options['state'] = 'readonly'
        options.current(0)
        self.state = True

        showinfo(title='Success!',
                 message='File successfully open')


class Create(Excel):
    def __init__(self):
        super().__init__()
        self.month_odoo = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', "Sep", 'Oct', 'Nov', 'Dec']
        self.summary_header = ['Odoo', 'Excel', 'Difference', 'Remarks']
        self.month_pc = []
        self.header_row_odoo = 4
        self.pc_sheet = None
        self.odoo_sheet = None
        self.max_col_odoo = None
        self.max_row_odoo = None
        self.max_col_pc = None
        self.max_row_pc = None
        self.cell_odoo = None
        self.cell_pc = None
        self.month_selected = None
        self.month_index = None
        self.selected_col = None

    def create_excel(self, app):
        self.workbook = Workbook()
        self.sheet_list = self.workbook.sheetnames
        self.pc_sheet = self.workbook.create_sheet(str(app.pc_options.get()))
        self.odoo_sheet = self.workbook.create_sheet(str(app.odoo_options.get()))

    def load_attribute(self):
        self.max_col_odoo = self.odoo_sheet.max_column
        self.max_row_odoo = self.odoo_sheet.max_row
        self.max_col_pc = self.pc_sheet.max_column
        self.max_row_pc = self.pc_sheet.max_row
        self.cell_odoo = self.odoo_sheet.cell
        self.cell_pc = self.pc_sheet.cell
        self.month_selected = self.cell_odoo(4, 7).value[-3:]
        self.month_index = self.month_odoo.index(self.month_selected)
