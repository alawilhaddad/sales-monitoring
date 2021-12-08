from openpyxl.utils import get_column_letter
import string
from excel.copy_sheet_data import copy_cell_style


def general(new):
    # Initiate variable
    new.load_attribute()

    # Hide unused column
    for col in range(1, new.max_col_odoo):
        empty = True
        for row in range(5, new.max_row_odoo):
            if new.cell_odoo(row, col).value != 0:
                empty = False
                break
        if empty:
            new.odoo_sheet.column_dimensions[get_column_letter(col)].hidden = True

    # Make smaller column for separator
    new.odoo_sheet.column_dimensions[get_column_letter(new.max_col_odoo+1)].width = 5

    # Freeze pane odoo
    new.odoo_sheet.freeze_panes = "C5"

    # Create summary table
    summary_start_index = new.max_col_odoo + 2
    summary_end_index = summary_start_index + len(new.summary_header) - 1
    for i, column in enumerate(range(summary_start_index, summary_end_index + 1)):
        new.cell_odoo(new.header_row_odoo, column).value = new.summary_header[i]
        new.odoo_sheet.column_dimensions[get_column_letter(column)].width = 20
        for row in range(4, new.max_row_odoo + 1):
            copy_cell_style(new.cell_odoo, row, column, row, new.max_col_odoo)

    # Sum Row
    for col in range(5, summary_end_index):
        if col != new.max_col_odoo + 1:
            letter = get_column_letter(col)
            new.cell_odoo(new.max_row_odoo,
                          col).value = f'=SUM({letter}5:{letter}{new.max_row_odoo-1})'

    # Reformat Value
    for col in range(5, summary_end_index):
        for row in range(5, new.max_row_odoo + 1):
            new.cell_odoo(row, col).number_format = '#,##0.00;[RED]- #,##0.00;-'

    # Odoo recap column
    column = new.max_col_odoo + 2
    for row in range(5, new.max_row_odoo):
        new.cell_odoo(row, column).value = f"=sum(G{row}:K{row})"


def eni(new):
    # Reformat Odoo sheet
    general(new)

    # Reformat COFF number
    for row in range(5, new.max_row_odoo):
        value = str(new.cell_odoo(row, 2).value)
        new.cell_odoo(row, 2).value = value[7:]
        data_type = True
        for char in value[7:]:
            if char in string.digits:
                data_type = True
            elif char == ' ':
                value.replace(' ', '')
            else:
                data_type = False
                break
        if data_type:
            new.cell_odoo(row, 2).number_format = '0'
            new.cell_odoo(row, 2).value = int(value[7:])
        else:
            pass

    # Search selected month on PC
    for col in range(1, new.max_col_pc):
        if new.cell_pc(7, col).value == new.month_pc[new.month_index]:
            new.selected_col = get_column_letter(col)

    # PC recap column
    column = new.max_col_odoo + 3
    for row in range(5, new.max_row_odoo):
        new.cell_odoo(row, column).value = f'=IFERROR(VLOOKUP(B{row},CHOOSE(' + '{1,2},' + \
                                       f'{new.workbook.sheetnames[0]}!$D$14:$D$800,{new.workbook.sheetnames[0]}!' \
                                       f'${new.selected_col}$14:${new.selected_col}$800),2,0),0)'

    # Difference column
    column = new.max_col_odoo + 4
    for row in range(5, new.max_row_odoo):
        new.cell_odoo(row, column).value = f'={new.cell_odoo(row, column - 2).coordinate}-' \
                                       f'{new.cell_odoo(row, column - 1).coordinate}'

    # Freeze panes PC
    new.pc_sheet.freeze_panes = "F9"


def phkt(new):
    # Reformat Odoo sheet
    general(new)

    # Search selected month on PC
    for col in range(1, new.max_col_pc):
        if new.cell_pc(1, col).value == new.month_pc[new.month_index]:
            new.selected_col = get_column_letter(col + 1)

    # PC recap column
    column = new.max_col_odoo + 3
    for row in range(5, new.max_row_odoo):
        new.cell_odoo(row, column).value = \
            f"=IFERROR(VLOOKUP(B{row},CHOOSE(" + '{1,2},' + \
            f"'{new.workbook.sheetnames[0]}'!$F$6:$F$800,'{new.workbook.sheetnames[0]}'!" \
            f"${new.selected_col}$6:${new.selected_col}$800),2,0),0)"

    # Difference column
    column = new.max_col_odoo + 4
    for row in range(5, new.max_row_odoo):
        new.cell_odoo(row, column).value = f'={new.cell_odoo(row, column - 2).coordinate}-' \
                                           f'{new.cell_odoo(row, column - 1).coordinate}'

    # Freeze panes PC
    new.pc_sheet.freeze_panes = "H6"


def phm_edi(new):
    # Reformat Odoo sheet
    general(new)

    # Search selected month on PC
    for col in range(1, new.max_col_pc):
        value = str(new.cell_pc(6, col).value)

        if value[:7] == new.month_pc[new.month_index]:
            new.selected_col = get_column_letter(col)

    # Reformat CE number odoo
    for row in range(5, new.max_row_odoo):
        value = str(new.cell_odoo(row, 2).value)
        if value.isnumeric():
            new.cell_odoo(row, 2).number_format = '0'
            new.cell_odoo(row, 2).value = int(value)

    # Reformat COFF number odoo
    for row in range(11, new.max_row_pc):
        value = str(new.cell_pc(row, 5).value)
        if value.isnumeric():
            new.cell_pc(row, 5).number_format = '0'
            new.cell_pc(row, 5).value = int(value)

    # PC recap column
    column = new.max_col_odoo + 3
    for row in range(5, new.max_row_odoo):
        new.cell_odoo(row, column).value = \
            f"=IFERROR(VLOOKUP(B{row},CHOOSE(" + '{1,2},' + \
            f"'{new.workbook.sheetnames[0]}'!$E$10:$E$1018,'{new.workbook.sheetnames[0]}'!" \
            f"${new.selected_col}$10:${new.selected_col}$1018),2,0),0)"

    # Difference column
    column = new.max_col_odoo + 4
    for row in range(5, new.max_row_odoo):
        new.cell_odoo(row, column).value = f'={new.cell_odoo(row, column - 2).coordinate}-' \
                                            f'{new.cell_odoo(row, column - 1).coordinate}'

    # Freeze panes PC
    new.pc_sheet.freeze_panes = "N11"
