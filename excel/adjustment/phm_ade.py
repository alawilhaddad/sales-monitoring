from excel.var import new, month_odoo, month_number, header
import string
from openpyxl.utils import get_column_letter
from copy import copy
import datetime


def phm_ade():
    max_col_odoo = new.odoo_sheet.max_column
    max_row_odoo = new.odoo_sheet.max_row
    max_col_pc = new.pc_sheet.max_column
    cell_odoo = new.odoo_sheet.cell
    cell_pc = new.pc_sheet.cell
    month = cell_odoo(4, 7).value[-3:]
    month_index = month_odoo.index(month)

    for col in range(1, max_col_pc):
        month_value = cell_pc(2, col).value
        if type(month_value) == datetime.datetime:
            if month_value.strftime("%m") == month_number[month_index]:
                selected_col = get_column_letter(col + 1)

    for col in range(2):
        # column = max_col_odoo + col
        for row in range(4, max_row_odoo + 1):
            if col == 0:
                if 4 < row < max_row_odoo:
                    value = str(cell_odoo(row, 2).value)
                    cell_odoo(row, 2).value = value[0:]
                    data_type = True
                    for char in value[0:]:
                        if char in string.digits:
                            data_type = True
                        elif char == ' ':
                            value.replace(' ', '')
                        else:
                            data_type = False
                            break
                    if data_type is True:
                        cell_odoo(row, 2).number_format = '0'
                        cell_odoo(row, 2).value = int(value[0:])
                    else:
                        pass

    new.odoo_sheet.insert_cols(3)
    new.odoo_sheet.merged_cells.remove(f'A{max_row_odoo}:D{max_row_odoo}')

    for row in range(4, max_row_odoo):
        cell_odoo(row, 3).value = cell_odoo(row, 2).value
        cell_odoo(row, 3).number_format = copy(cell_odoo(row, 2).number_format)
        cell_odoo(row, 3).font = copy(cell_odoo(row, 2).font)
        cell_odoo(row, 3).border = copy(cell_odoo(row, 2).border)
        cell_odoo(row, 3).fill = copy(cell_odoo(row, 2).fill)
        cell_odoo(row, 3).protection = copy(cell_odoo(row, 2).protection)
        cell_odoo(row, 3).alignment = copy(cell_odoo(row, 2).alignment)

    cell_odoo(4, 2).value = 'No. CE Odoo'
    cell_odoo(4, 2).value = 'No. CE Koreksi'
    new.odoo_sheet.merge_cells(f'A{max_row_odoo}:E{max_row_odoo}')

    new.odoo_sheet.column_dimensions[get_column_letter(max_col_odoo + 2)].width = 5

    for col in range(4):
        column = max_col_odoo + col + 3
        for row in range(4, max_row_odoo + 1):
            if row == 4:
                cell_odoo(row, column).value = header[col]
            elif 4 < row < max_row_odoo:
                if col == 0:
                    cell_odoo(row, column).value = f"=sum(H{row}:L{row})"
                elif col == 1:
                    cell_odoo(row,
                              column).value = f'=IFERROR(VLOOKUP(C{row},CHOOSE(' + '{1,2},' + \
                                              f'{new.wb.sheetnames[0]}!$B$14:$B$800,' \
                                              f'{new.wb.sheetnames[0]}!${selected_col}$14:${selected_col}$800),2,0),0)'
                elif col == 2:
                    cell_odoo(row,
                              column).value = f'={cell_odoo(row, column - 2).coordinate}-' \
                                              f'{cell_odoo(row, column - 1).coordinate}'
            elif row == max_row_odoo:
                if col == 3:
                    pass
                else:
                    loc = get_column_letter(column)
                    cell_odoo(row, column).value = f"=sum({loc}5:{loc}{max_row_odoo - 1})"
            new.odoo_sheet.column_dimensions[get_column_letter(column)].width = 20
            cell_odoo(row, column).number_format = copy(cell_odoo(row, max_col_odoo).number_format)
            cell_odoo(row, column).font = copy(cell_odoo(row, max_col_odoo).font)
            cell_odoo(row, column).border = copy(cell_odoo(row, max_col_odoo).border)
            cell_odoo(row, column).fill = copy(cell_odoo(row, max_col_odoo).fill)
            cell_odoo(row, column).protection = copy(cell_odoo(row, max_col_odoo).protection)
            cell_odoo(row, column).alignment = copy(cell_odoo(row, max_col_odoo).alignment)
    max_col_odoo = new.odoo_sheet.max_column

    for col in range(8, 13):
        empty = True
        for row in range(5, max_row_odoo):
            if cell_odoo(row, col).value != 0:
                empty = False
                break
        if empty:
            new.odoo_sheet.column_dimensions[get_column_letter(col)].hidden = True

    for col in range(6, 14):
        cell_odoo(max_row_odoo,
                  col).value = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{max_row_odoo - 1})'

    for col in range(6, max_col_odoo + 5):
        for row in range(5, max_row_odoo + 1):
            cell_odoo(row, col).number_format = '#,##0.00;[RED]- #,##0.00;-'
