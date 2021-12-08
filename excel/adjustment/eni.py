from excel.var import new, pc, month_odoo
import string
from openpyxl.utils import get_column_letter
from copy import copy


def eni(new, pc):
    pc.month = ["JAN", "FEB", "MAR", "APR", "MEI", "JUNI", "JULI", "AGST", "SEPT", "OKT", "NOV", "DES"]
    max_col_odoo = new.odoo_sheet.max_column
    max_row_odoo = new.odoo_sheet.max_row
    max_col_pc = new.pc_sheet.max_column
    cell_odoo = new.odoo_sheet.cell
    cell_pc = new.pc_sheet.cell
    month = cell_odoo(4, 7).value[-3:]
    month_index = month_odoo.index(month)

    for col in range(1, max_col_pc):
        if cell_pc(7, col).value == pc.month[month_index]:
            selected_col = get_column_letter(col)

    # reworks default form odoo
    for col in range(2):
        column = max_col_odoo + col
        for row in range(4, max_row_odoo + 1):
            if col == 0:
                if 4 < row < max_row_odoo:
                    value = str(cell_odoo(row, 2).value)
                    cell_odoo(row, 2).value = value[7:]
                    data_type = True
                    for char in value[7:]:
                        if char in string.digits:
                            data_type = True
                        elif char == ' ':
                            value.replace(' ', '')
                        else:
                            data_type = False
                            break
                    if data_type is True:
                        cell_odoo(row, 2).number_format = '0'
                        cell_odoo(row, 2).value = int(value[7:])
                    else:
                        pass
            elif col == 1:
                new.odoo_sheet.column_dimensions[get_column_letter(column)].width = 5
    # create summary
    for col in range(4):
        column = max_col_odoo + col + 2
        for row in range(4, max_row_odoo + 1):
            if row == 4:
                cell_odoo(row, column).value = header[col]
            elif 4 < row < max_row_odoo:
                if col == 0:
                    cell_odoo(row, column).value = f"=sum(G{row}:K{row})"
                elif col == 1:
                    cell_odoo(row,
                              column).value = f'=IFERROR(VLOOKUP(B{row},CHOOSE(' + '{1,2},' + \
                                              f'{new.workbook.sheetnames[0]}!$D$14:$D$800,{new.workbook.sheetnames[0]}!'\
                                              f'${selected_col}$14:${selected_col}$800),2,0),0)'
                elif col == 2:
                    cell_odoo(row,
                              column).value = f'={cell_odoo(row, column - 2).coordinate}-' \
                                              f'{cell_odoo(row, column - 1).coordinate}'
            elif row == max_row_odoo:
                if col == 3:
                    pass
                else:
                    loc = get_column_letter(column)
                    cell_odoo(row, column).value = f"=sum({loc}5:{loc}{max_row_odoo - 1})"
            new.odoo_sheet.column_dimensions[get_column_letter(column)].width = 20
            cell_odoo(row, column).number_format = copy(cell_odoo(row, max_col_odoo).number_format)
            cell_odoo(row, column).font = copy(cell_odoo(row, max_col_odoo).font)
            cell_odoo(row, column).border = copy(cell_odoo(row, max_col_odoo).border)
            cell_odoo(row, column).fill = copy(cell_odoo(row, max_col_odoo).fill)
            cell_odoo(row, column).protection = copy(cell_odoo(row, max_col_odoo).protection)
            cell_odoo(row, column).alignment = copy(cell_odoo(row, max_col_odoo).alignment)

    for col in range(7, 12):
        empty = True
        for row in range(5, max_row_odoo):
            if cell_odoo(row, col).value != 0:
                empty = False
                break
        if empty:
            new.odoo_sheet.column_dimensions[get_column_letter(col)].hidden = True

    for col in range(5, 13):
        cell_odoo(max_row_odoo,
                  col).value = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{max_row_odoo - 1})'

    for col in range(5, max_col_odoo + 5):
        for row in range(5, max_row_odoo + 1):
            cell_odoo(row, col).number_format = '#,##0.00;[RED]- #,##0.00;-'
