from excel.var import new, pc, month_odoo, header
from openpyxl.utils import get_column_letter
from copy import copy


def phkt():
    pc.month = ['JANUARI 2021', 'FEBRUARI 2021', 'MARET 2021', 'APRIL 2021', 'MEI 2021', 'JUNI 2021',
                'JULI 2021', 'AGUSTUS 2021', 'SEPTEMBER 2021', 'OKTOBER 2021', 'NOVEMBER 2021', 'DESEMBER 2021']
    max_col_odoo = new.odoo_sheet.max_column
    max_row_odoo = new.odoo_sheet.max_row
    max_col_pc = new.pc_sheet.max_column
    cell_odoo = new.odoo_sheet.cell
    cell_pc = new.pc_sheet.cell
    month = cell_odoo(4, 7).value[-3:]
    month_index = month_odoo.index(month)

    for col in range(1, max_col_pc):
        if cell_pc(1, col).value == pc.month[month_index]:
            selected_col = get_column_letter(col + 1)

    new.odoo_sheet.column_dimensions[get_column_letter(max_col_odoo + 1)].width = 5

    for col in range(4):
        column = max_col_odoo + col + 2
        for row in range(4, max_row_odoo + 1):
            if row == 4:
                cell_odoo(row, column).value = header[col]
            elif 4 < row < max_row_odoo:
                if col == 0:
                    cell_odoo(row, column).value = f"=sum(G{row}:K{row})"
                elif col == 1:
                    cell_odoo(row,
                              column).value = f"=IFERROR(VLOOKUP(B{row},CHOOSE(" + '{1,2},' + \
                                              f"'{new.wb.sheetnames[0]}'!$F$6:$F$800,'{new.wb.sheetnames[0]}'!" \
                                              f"${selected_col}$6:${selected_col}$800),2,0),0)"
                elif col == 2:
                    cell_odoo(row,
                              column).value = f'={cell_odoo(row, column - 2).coordinate}-' \
                                              f'{cell_odoo(row, column - 1).coordinate}'
            elif row == max_row_odoo:
                if col == 3:
                    pass
                else:
                    loc = get_column_letter(column)
                    cell_odoo(row, column).value = f"=sum({loc}5:{loc}{max_row_odoo - 1})"
            new.odoo_sheet.column_dimensions[get_column_letter(column)].width = 20
            cell_odoo(row, column).number_format = copy(cell_odoo(row, max_col_odoo).number_format)
            cell_odoo(row, column).font = copy(cell_odoo(row, max_col_odoo).font)
            cell_odoo(row, column).border = copy(cell_odoo(row, max_col_odoo).border)
            cell_odoo(row, column).fill = copy(cell_odoo(row, max_col_odoo).fill)
            cell_odoo(row, column).protection = copy(cell_odoo(row, max_col_odoo).protection)
            cell_odoo(row, column).alignment = copy(cell_odoo(row, max_col_odoo).alignment)
    max_col_odoo = new.odoo_sheet.max_column

    for col in range(8, 13):
        empty = True
        for row in range(5, max_row_odoo):
            if cell_odoo(row, col).value != 0:
                empty = False
                break
        if empty:
            new.odoo_sheet.column_dimensions[get_column_letter(col)].hidden = True

    for col in range(5, 13):
        cell_odoo(max_row_odoo,
                  col).value = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{max_row_odoo - 1})'

    for col in range(5, max_col_odoo + 5):
        for row in range(5, max_row_odoo + 1):
            cell_odoo(row, col).number_format = '#,##0.00;[RED]- #,##0.00;-'
