import sys
from datetime import *
from excel.adjust import eni, phkt, phm_edi
from excel.copy_excel import *
from excel.helper import *
from excel import var
from tkinter.filedialog import asksaveasfilename, askdirectory
from tkinter.messagebox import showinfo, showwarning, askyesno
import openpyxl as ex
from os import startfile
import pywintypes
from zipfile import BadZipfile


def exit_app():
    sys.exit()
    quit()



def home_show(self, pc, odoo):
    # Hide other menu
    setting_hide(self)
    guide_hide(self)

    # Main title
    self.canvas.create_image(
        140, 92,
        anchor="nw",
        image=self.main_title_img,
        tags="home")

    # Background for file name pc
    self.canvas.create_image(
        140, 310,
        anchor="nw",
        image=self.label_canvas,
        tags="home")

    # Background for file name odoo
    self.canvas.create_image(
        140, 390,
        anchor="nw",
        image=self.label_canvas,
        tags="home")

    # Upper blue rectangle
    self.canvas.create_rectangle(
        370, 270, 370 + 200, 270 + 40,
        fill="#81c0f2",
        outline="",
        tags="home")

    # Lower blue rectangle
    self.canvas.create_rectangle(
        370, 350, 370 + 200, 350 + 40,
        fill="#81c0f2",
        outline="",
        tags="home")

    # PC file name
    self.pc_label = self.canvas.create_text(
        150.0, 322.0,
        fill="#ffffff",
        anchor="w",
        font=("KarlaTamilUpright-Regular", 9),
        tags="home")
    self.canvas.itemconfigure(self.pc_label, text=pc.filename)

    # Odoo file name
    self.odoo_label = self.canvas.create_text(
        150.0, 401.0,
        fill="#ffffff",
        anchor="w",
        font=("KarlaTamilUpright-Regular", 9),
        tags="home")
    self.canvas.itemconfigure(self.odoo_label, text=odoo.filename)

    # Load PC button
    self.pc_button.place(
        x=140, y=270,
        width=230,
        height=40)

    # PC sheet options
    self.pc_options.place(x=610, y=270, width=150, height=40)

    # Password title
    self.canvas.create_text(
        395.0, 280.0,
        text="Pass: ",
        fill="#ffffff",
        anchor="nw",
        font=("KarlaTamilUpright-Bold", 11),
        tags="home")

    # Entry box for password
    self.canvas.create_image(
        440.0, 275.0,
        anchor="nw",
        image=self.entry0_img,
        tags="home")
    self.entry0.place(
        x=445.0, y=277,
        anchor="nw",
        width=90,
        height=28)

    # Load reload button for PC
    self.reload_pc_button.place(
        x=570, y=270,
        width=40,
        height=40)

    # Load Odoo button
    self.odoo_button.place(
        x=140, y=350,
        width=230,
        height=40)

    # Odoo sheet options
    self.odoo_options.place(x=610, y=350, width=150, height=40)

    # Load reload button for Odoo
    self.reload_odoo_button.place(
        x=570, y=350,
        width=40,
        height=40)

    # Start button
    self.start_button.place(
        x=610, y=460,
        width=150,
        height=40)


def home_hide(self):
    # Clear object with 'home' tag
    self.canvas.delete("home")

    # Clear button in home section
    self.pc_button.place_forget()
    self.pc_options.place_forget()
    self.reload_pc_button.place_forget()
    self.entry0.place_forget()
    self.odoo_button.place_forget()
    self.reload_odoo_button.place_forget()
    self.odoo_options.place_forget()
    self.start_button.place_forget()


def guide_show(self):
    # Clear object without 'guide' tag
    home_hide(self)
    setting_hide(self)

    # Guide title
    self.canvas.create_image(
        140, 96,
        anchor="nw",
        image=self.help_title_img,
        tags="guide")

    # Guide content
    self.canvas.create_text(
        150.0, 210.0,
        text="1. Download form monitoring yang dikirimkan Project Controller\n"
             "2. Download sales report dari Odoo, pastikan tanggal yang dipilih\n"
             "    mencakup 1 bulan penuh\n"
             '3. Klik tombol "Monitoring by PC", pilih file yang diinginkan\n'
             '4. Klik tombol "Monitoring Odoo", pilih file yang diingingkan\n'
             "5. Pilih sheet monitoring yang diperlukan untuk masing-masing file\n"
             '6. Klik "Start"\n'
             "7. Simpan file output sesuai format yang diinginkan\n"
             "8. Adjust file output sesuai kebutuhan\n"
             "9. Selesai",
        fill="#ffffff",
        anchor="nw",
        font=("Cabin-Regular", int(14.0)),
        tags="guide")


def guide_hide(self):
    # Hide object with "guide" tag
    self.canvas.delete("guide")


def setting_show(self):
    # Hide object without "setting" tag
    home_hide(self)
    guide_hide(self)

    # Setting title
    self.canvas.create_image(
        140, 96,
        anchor="nw",
        image=self.setting_title_img,
        tags="setting")

    # Open directory background
    self.canvas.create_image(
        140, 310,
        anchor="nw",
        image=self.label_round,
        tags="setting")

    # Open directory label
    self.open_label = self.canvas.create_text(
        150.0, 322.0,
        fill="#ffffff",
        anchor="w",
        font=("KarlaTamilUpright-Regular", 9),
        tags="setting")
    self.canvas.itemconfigure(self.open_label, text=var.default_open_dir)

    # Save directory background
    self.canvas.create_image(
        140, 390,
        anchor="nw",
        image=self.label_round,
        tags="setting")

    # Save directory label
    self.save_label = self.canvas.create_text(
        150.0, 401.0,
        fill="#ffffff",
        anchor="w",
        font=("KarlaTamilUpright-Regular", 9),
        tags="setting")
    self.canvas.itemconfigure(self.save_label, text=var.default_save_dir)

    # Set open directory button
    self.open_d_button.place(
        x=140, y=270,
        width=230,
        height=40)

    # Set save directory button
    self.save_d_button.place(
        x=140, y=350,
        width=230,
        height=40)


def setting_hide(self):
    # Hide object with "setting" tag
    self.canvas.delete('setting')

    # Hide object setting button
    self.open_d_button.place_forget()
    self.save_d_button.place_forget()


def open_excel(app, main_source, aux_source, options, label):
    try:
        main_source.load(app, options, label)
    except BadZipfile:
        try:
            main_source.path = unlock_excel(main_source.path, str(app.password.get()))
            main_source.workbook = px.load_workbook(main_source.path, data_only=True)
            main_source.load_attribute(app, options, label)
        #
        except pywintypes.com_error:
            showwarning(title='Incorrect Password!',
                        message='Please make sure you type in correct Password')

    if main_source.state and aux_source.state:
        app.start_button.config(state="normal")


def reload(main, source, options, label):
    if askyesno(title="Clear Item", message="Are you sure you want to clear this section"):
        clear(main, source, options, label)
    else:
        pass


def clear(main, source, options, label):
    # Clear selection
    source.path = None
    source.filename = None
    source.workbook = None
    source.sheet_list = []
    source.active_sheet = None
    source.state = False
    source.image = None
    source.options = None
    source.var = None
    source.cell = None
    source.month = []
    source.label = None
    source.selected_sheet = None
    main.canvas.itemconfigure(label, text='')
    options['values'] = ['']
    options['state'] = 'disable'
    options.current(0)
    source.state = True
    main.start_button['state'] = 'disable'


def start(app, new_excel, pc_excel, odoo_excel):
    new_excel.create_excel(app)

    # Copy sheet from pc and odoo to new workbook
    pc_excel.active_sheet = pc_excel.workbook[app.pc_options.get()]
    copy_sheet(pc_excel.active_sheet, new_excel.pc_sheet)
    odoo_excel.active_sheet = odoo_excel.workbook[app.odoo_options.get()]
    copy_sheet(odoo_excel.active_sheet, new_excel.odoo_sheet)

    # Determine project using project number in odoo's sheet name
    if 'Sheet' in new_excel.sheet_list:
        new_excel.workbook.remove(new_excel.workbook['Sheet'])
    if new_excel.workbook.sheetnames[1] == '3250':
        new_excel.month_pc = ["JAN", "FEB", "MAR", "APR", "MEI", "JUNI", "JULI", "AGST", "SEPT", "OKT", "NOV", "DES"]
        eni(new_excel)
    elif new_excel.workbook.sheetnames[1] == '3235':
        new_excel.month_pc = ["2021-01", "2021-02", "2021-03", "2021-04", "2021-05", "2021-06", "2021-07", "2021-08",
                              "2021-09", "2021-10", "2021-11", "2021-12"]
        phm_edi(new_excel)
    elif new_excel.workbook.sheetnames[1] == '3247':
        new_excel.month_pc = ['JANUARI 2021', 'FEBRUARI 2021', 'MARET 2021', 'APRIL 2021', 'MEI 2021',
                              'JUNI 2021', 'JULI 2021', 'AGUSTUS 2021', 'SEPTEMBER 2021', 'OKTOBER 2021',
                              'NOVEMBER 2021', 'DESEMBER 2021']
        phkt(new_excel)

    # Set time and week
    dates = datetime.now()
    date_number = int(dates.strftime("%d"))
    month = dates.strftime("%m")
    year = dates.strftime("%Y")
    project_number = new_excel.workbook.sheetnames[1]
    if date_number < 8:
        week = "01"
    elif 7 < date_number < 15:
        week = "02"
    elif 14 < date_number < 22:
        week = "03"
    elif 21 < date_number:
        week = "04"

    # Save as new file, odoo file, and pc file
    path = asksaveasfilename(
        initialdir=var.default_save_dir,
        title='Save Validation File',
        initialfile=f'{year}-{month}W{week}-Validation-{project_number}-v01.xlsx',
        filetypes=(("Excel File", "*xlsx"), ("All Files", "*.*")))
    if path != '':
        save_dir = '/'.join(path.split("/")[:-1])
        new_excel.workbook.save(path)
        pc_excel.workbook.save(asksaveasfilename(
                initialdir=save_dir,
                title='Save PC File',
                initialfile=f'{year}-{month}W{week}-PC-{project_number}-v01.xlsx',
                filetypes=(("Excel File", "*xlsx"), ("All Files", "*.*"))))
        odoo_excel.workbook.save(asksaveasfilename(
                initialdir=save_dir,
                title='Save Odoo File',
                initialfile=f'{year}-{month}W{week}-Odoo-{project_number}-v01.xlsx',
                filetypes=(("Excel File", "*xlsx"), ("All Files", "*.*"))))
        showinfo(title='Done', message='Done!')
        startfile(path)
        clear(app, pc_excel, app.pc_options, app.pc_label)
        clear(app, odoo_excel, app.odoo_options, app.odoo_label)
    else:
        return


def default_open(app):
    # Set default open directory
    var.default_open_dir = askdirectory(initialdir=var.default_open_dir,title="Default Open Directory")
    if var.default_open_dir:
        with open("config.txt", "r+") as file:
            lines = file.readlines()
            lines[0] = f"open={var.default_open_dir}\n"
        with open("config.txt", "w") as file:
            file.writelines(lines)
        app.canvas.itemconfigure(app.open_label, text=var.default_open_dir)

def default_save(app):
    # Set default save directory
    var.default_save_dir = askdirectory(initialdir=var.default_save_dir,title="Default Save Directory")
    if var.default_save_dir:
        with open("config.txt", "r+") as file:
            lines = file.readlines()
            lines[1] = f"save={var.default_save_dir}"
        with open("config.txt", "w") as file:
            file.writelines(lines)
        app.canvas.itemconfigure(app.save_label, text=var.default_save_dir)
