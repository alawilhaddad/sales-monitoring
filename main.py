from copy import copy
from zipfile import *
from tkinter import *
from tkinter import ttk, Toplevel
from tkinter import filedialog
from tkinter import messagebox
from openpyxl.utils import get_column_letter
from math import ceil
import openpyxl
import string
import os
import datetime
import win32com.client


class Excel:
    def __init__(self):
        self.path = ''
        self.wb = ''
        self.sheet_list = ''
        self.active_sheet = ''
        self.state = False
        self.temp_list = ['']


class Load(Excel):
    def __init__(self, _window, _canvas):
        super().__init__()
        self.button = Button(
            borderwidth=0,
            highlightthickness=0,
            relief="flat")
        self.button['command'] = lambda source=self: _open(source)
        self.label = canvas.create_text(0, 0, text='')
        self.filename = ''
        self.var = StringVar(window)
        self.var.set(self.temp_list[0])
        self.options = ttk.Combobox(
            window,
            textvariable=self.var,
            width=12,
            font=('karla tamil upright', 9),
            justify='right',
            state="readonly")
        self.options['state'] = 'disabled'

    def set_excel(self):
        self.path = filedialog.askopenfilename(initialdir='C:/Users/mohin/Downloads/',
                                               title='Open File',
                                               filetypes=(("Excel File", "*xlsx"), ("All Files", "*.*")))

        self.filename = self.path.split('/')[-1]

        self.path = convert_slash(self.path)
        if self.path is None:
            return
        elif self.path != '':
            try:
                self.wb = openpyxl.load_workbook(self.path, data_only=True)
                self.sheet_list = self.wb.sheetnames
                self.active_sheet = self.sheet_list[0]
                canvas.itemconfigure(self.label, text=self.filename)
                self.options['values'] = self.sheet_list
                self.options['state'] = 'readonly'
                self.var.set(self.sheet_list[0])
                self.state = True
            except BadZipfile:
                messagebox.showinfo(title='Password Protected', message='Your file is password protected.')
                self.path = remove_password_xlsx(self.path, input("password: "))
                self.filename = self.path.split('\\')[-1]
                self.wb = openpyxl.load_workbook(self.path, data_only=True)
                self.sheet_list = self.wb.sheetnames
                self.active_sheet = self.sheet_list[0]
                canvas.itemconfigure(self.label, text=self.filename)
                self.options['values'] = self.sheet_list
                self.options['state'] = 'readonly'
                self.var.set(self.sheet_list[0])
                self.state = True
            start_state()

        else:
            return


class Create(Excel):
    def __init__(self):
        super().__init__()
        self.pc_sheet = None
        self.odoo_sheet = None

    def create_excel(self):
        self.wb = openpyxl.Workbook()
        self.sheet_list = self.wb.sheetnames
        self.pc_sheet = self.wb.create_sheet(pc.var.get())
        self.odoo_sheet = self.wb.create_sheet(odoo.var.get())


class Win(Tk):
    def __init__(self):
        super().__init__()
        super().overrideredirect(True)
        self.offset_x = 0
        self.offset_y = 0
        super().bind("<Button-1>", self.click_win)
        super().bind("<B1-Motion>", self.drag_win)

    def drag_win(self, event):
        psx = super().winfo_pointerx() - self.offset_x
        psy = super().winfo_pointery() - self.offset_y
        super().geometry(f"+{psx}+{psy}")

    def click_win(self, event):
        self.offset_x = super().winfo_pointerx() - super().winfo_rootx()
        self.offset_y = super().winfo_pointery() - super().winfo_rooty()


def convert_slash(path):
    path_list = list(path)
    for char in path_list:
        if char == '/':
            index = path_list.index(char)
            path_list[index] = '\\'
    path = ''.join(path_list)
    return path


def remove_password_xlsx(filename, pw_str):
    xcl = win32com.client.Dispatch("Excel.Application")
    wb = xcl.Workbooks.Open(filename, False, False, None, pw_str)
    xcl.DisplayAlerts = False
    filename_split = filename.split('.')
    filename_split[0] += '_unlocked'
    filename = '.'.join(filename_split)
    wb.SaveAs(filename, None, '', '')
    xcl.Quit()
    return filename


def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        pass
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width)
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


def _open(source):
    source.set_excel()


def quit_help():
    help_window.destroy()
    window.deiconify()


def _help():
    global help_window
    window.withdraw()
    help_window = Toplevel()
    help_window.configure(bg="#ffffff")
    help_window.overrideredirect(True)
    help_width = help_window.winfo_screenwidth()
    help_height = help_window.winfo_screenheight()
    help_position_x = (help_width / 2) - 200
    help_position_y = (help_height / 2) - 150
    help_window.geometry(f"400x300+{int(help_position_x)}+{int(help_position_y)}")
    help_window.overrideredirect(True)
    help_canvas = Canvas(
        help_window,
        bg="#ffffff",
        height=300,
        width=400,
        bd=0,
        highlightthickness=0,
        relief="ridge")
    help_canvas.place(x=0, y=0)

    help_background_img = PhotoImage(file=f"win/img/help_background.png")
    help_canvas.create_image(
        200.0, 150.0,
        image=help_background_img)

    img0 = PhotoImage(file=f"win/img/help_close.png")
    b0 = Button(
        help_window,
        image=img0,
        borderwidth=0,
        highlightthickness=0,
        command=quit_help,
        relief="flat")

    b0.place(
        x=375, y=7,
        width=19,
        height=19)

    window.resizable(False, False)
    help_window.mainloop()


def start():
    global project
    new.create_excel()

    pc.active_sheet = pc.wb[pc.var.get()]
    copy_sheet(pc.active_sheet, new.pc_sheet)
    odoo.active_sheet = odoo.wb[odoo.var.get()]
    copy_sheet(odoo.active_sheet, new.odoo_sheet)
    if 'Sheet' in new.sheet_list:
        new.wb.remove(new.wb['Sheet'])
    if new.wb.sheetnames[1] == '3250':
        eni_adjustment()
        project = 'ENI'
    elif new.wb.sheetnames[1] == '3235':
        phm_adjustment()
        project = 'PHM'
    elif new.wb.sheetnames[1] == '3247':
        phkt_adjustment()
        project = 'PHKT'
    date = datetime.datetime.now()

    path = filedialog.asksaveasfilename(
        initialdir='C:/Users/mohin/Downloads/',
        title='Save File',
        initialfile=f'{project}-{new.wb.sheetnames[1]}-{date.strftime("%Y%m%d")}-0{ceil(float(date.strftime("%d"))//7)}\
        -Monitoring-v01.xlsx',
        filetypes=(("Excel File", "*xlsx"), ("All Files", "*.*")))
    if path == '':
        return
    else:
        new.wb.save(path)
        messagebox.showinfo(title='Done', message='Done!')
        os.startfile(path)
    quit()


def start_state():
    if pc.state and odoo.state:
        start_button.config(state=NORMAL)


def eni_adjustment():
    max_col_odoo = new.odoo_sheet.max_column
    max_row_odoo = new.odoo_sheet.max_row
    max_col_pc = new.pc_sheet.max_column
    cell_odoo = new.odoo_sheet.cell
    cell_pc = new.pc_sheet.cell
    month = cell_odoo(4, 7).value[-3:]
    month_index = month_odoo.index(month)

    for col in range(1, max_col_pc):
        if cell_pc(7, col).value == month_eni[month_index]:
            selected_col = get_column_letter(col)

    # reworks default form odoo
    for col in range(2):
        column = max_col_odoo + col
        for row in range(4, max_row_odoo + 1):
            if col == 0:
                if 4 < row < max_row_odoo:
                    value = str(cell_odoo(row, 2).value)
                    cell_odoo(row, 2).value = value[7:]
                    data_type = True
                    for char in value[7:]:
                        if char in string.digits:
                            data_type = True
                        elif char == ' ':
                            value.replace(' ', '')
                        else:
                            data_type = False
                            break
                    if data_type is True:
                        cell_odoo(row, 2).number_format = '0'
                        cell_odoo(row, 2).value = int(value[7:])
                    else:
                        pass
            elif col == 1:
                new.odoo_sheet.column_dimensions[get_column_letter(column)].width = 5
    # create summary
    for col in range(4):
        column = max_col_odoo + col + 2
        for row in range(4, max_row_odoo + 1):
            if row == 4:
                cell_odoo(row, column).value = header[col]
            elif 4 < row < max_row_odoo:
                if col == 0:
                    cell_odoo(row, column).value = f"=sum(G{row}:K{row})"
                elif col == 1:
                    cell_odoo(row,
                              column).value = f'=IFERROR(VLOOKUP(B{row},CHOOSE(' + '{1,2},' + \
                                              f'{new.wb.sheetnames[0]}!$D$14:$D$800,{new.wb.sheetnames[0]}!' \
                                              f'${selected_col}$14:${selected_col}$800),2,0),0)'
                elif col == 2:
                    cell_odoo(row,
                              column).value = f'={cell_odoo(row, column - 2).coordinate}-' \
                                              f'{cell_odoo(row, column - 1).coordinate}'
            elif row == max_row_odoo:
                if col == 3:
                    pass
                else:
                    loc = get_column_letter(column)
                    cell_odoo(row, column).value = f"=sum({loc}5:{loc}{max_row_odoo - 1})"
            new.odoo_sheet.column_dimensions[get_column_letter(column)].width = 20
            cell_odoo(row, column).number_format = copy(cell_odoo(row, max_col_odoo).number_format)
            cell_odoo(row, column).font = copy(cell_odoo(row, max_col_odoo).font)
            cell_odoo(row, column).border = copy(cell_odoo(row, max_col_odoo).border)
            cell_odoo(row, column).fill = copy(cell_odoo(row, max_col_odoo).fill)
            cell_odoo(row, column).protection = copy(cell_odoo(row, max_col_odoo).protection)
            cell_odoo(row, column).alignment = copy(cell_odoo(row, max_col_odoo).alignment)

    for col in range(7, 12):
        empty = True
        for row in range(5, max_row_odoo):
            if cell_odoo(row, col).value != 0:
                empty = False
                break
        if empty:
            new.odoo_sheet.column_dimensions[get_column_letter(col)].hidden = True

    for col in range(5, 13):
        cell_odoo(max_row_odoo,
                  col).value = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{max_row_odoo - 1})'

    for col in range(5, max_col_odoo + 5):
        for row in range(5, max_row_odoo + 1):
            cell_odoo(row, col).number_format = '#,##0.00;[RED]- #,##0.00;-'


def phm_adjustment():
    max_col_odoo = new.odoo_sheet.max_column
    max_row_odoo = new.odoo_sheet.max_row
    max_col_pc = new.pc_sheet.max_column
    cell_odoo = new.odoo_sheet.cell
    cell_pc = new.pc_sheet.cell
    month = cell_odoo(4, 7).value[-3:]
    month_index = month_odoo.index(month)

    for col in range(1, max_col_pc):
        month_value = cell_pc(2, col).value
        if type(month_value) == datetime.datetime:
            if month_value.strftime("%m") == month_number[month_index]:
                selected_col = get_column_letter(col + 1)

    for col in range(2):
        # column = max_col_odoo + col
        for row in range(4, max_row_odoo + 1):
            if col == 0:
                if 4 < row < max_row_odoo:
                    value = str(cell_odoo(row, 2).value)
                    cell_odoo(row, 2).value = value[0:]
                    data_type = True
                    for char in value[0:]:
                        if char in string.digits:
                            data_type = True
                        elif char == ' ':
                            value.replace(' ', '')
                        else:
                            data_type = False
                            break
                    if data_type is True:
                        cell_odoo(row, 2).number_format = '0'
                        cell_odoo(row, 2).value = int(value[0:])
                    else:
                        pass

    new.odoo_sheet.insert_cols(3)
    new.odoo_sheet.merged_cells.remove(f'A{max_row_odoo}:D{max_row_odoo}')

    for row in range(4, max_row_odoo):
        cell_odoo(row, 3).value = cell_odoo(row, 2).value
        cell_odoo(row, 3).number_format = copy(cell_odoo(row, 2).number_format)
        cell_odoo(row, 3).font = copy(cell_odoo(row, 2).font)
        cell_odoo(row, 3).border = copy(cell_odoo(row, 2).border)
        cell_odoo(row, 3).fill = copy(cell_odoo(row, 2).fill)
        cell_odoo(row, 3).protection = copy(cell_odoo(row, 2).protection)
        cell_odoo(row, 3).alignment = copy(cell_odoo(row, 2).alignment)

    cell_odoo(4, 2).value = 'No. CE Odoo'
    cell_odoo(4, 2).value = 'No. CE Koreksi'
    new.odoo_sheet.merge_cells(f'A{max_row_odoo}:E{max_row_odoo}')

    new.odoo_sheet.column_dimensions[get_column_letter(max_col_odoo + 2)].width = 5

    for col in range(4):
        column = max_col_odoo + col + 3
        for row in range(4, max_row_odoo + 1):
            if row == 4:
                cell_odoo(row, column).value = header[col]
            elif 4 < row < max_row_odoo:
                if col == 0:
                    cell_odoo(row, column).value = f"=sum(H{row}:L{row})"
                elif col == 1:
                    cell_odoo(row,
                              column).value = f'=IFERROR(VLOOKUP(C{row},CHOOSE(' + '{1,2},' + \
                                              f'{new.wb.sheetnames[0]}!$B$14:$B$800,' \
                                              f'{new.wb.sheetnames[0]}!${selected_col}$14:${selected_col}$800),2,0),0)'
                elif col == 2:
                    cell_odoo(row,
                              column).value = f'={cell_odoo(row, column - 2).coordinate}-' \
                                              f'{cell_odoo(row, column - 1).coordinate}'
            elif row == max_row_odoo:
                if col == 3:
                    pass
                else:
                    loc = get_column_letter(column)
                    cell_odoo(row, column).value = f"=sum({loc}5:{loc}{max_row_odoo - 1})"
            new.odoo_sheet.column_dimensions[get_column_letter(column)].width = 20
            cell_odoo(row, column).number_format = copy(cell_odoo(row, max_col_odoo).number_format)
            cell_odoo(row, column).font = copy(cell_odoo(row, max_col_odoo).font)
            cell_odoo(row, column).border = copy(cell_odoo(row, max_col_odoo).border)
            cell_odoo(row, column).fill = copy(cell_odoo(row, max_col_odoo).fill)
            cell_odoo(row, column).protection = copy(cell_odoo(row, max_col_odoo).protection)
            cell_odoo(row, column).alignment = copy(cell_odoo(row, max_col_odoo).alignment)
    max_col_odoo = new.odoo_sheet.max_column

    for col in range(8, 13):
        empty = True
        for row in range(5, max_row_odoo):
            if cell_odoo(row, col).value != 0:
                empty = False
                break
        if empty:
            new.odoo_sheet.column_dimensions[get_column_letter(col)].hidden = True

    for col in range(6, 14):
        cell_odoo(max_row_odoo,
                  col).value = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{max_row_odoo - 1})'

    for col in range(6, max_col_odoo + 5):
        for row in range(5, max_row_odoo + 1):
            cell_odoo(row, col).number_format = '#,##0.00;[RED]- #,##0.00;-'


def phkt_adjustment():
    global selected_col
    max_col_odoo = new.odoo_sheet.max_column
    max_row_odoo = new.odoo_sheet.max_row
    max_col_pc = new.pc_sheet.max_column
    cell_odoo = new.odoo_sheet.cell
    cell_pc = new.pc_sheet.cell
    month = cell_odoo(4, 7).value[-3:]
    month_index = month_odoo.index(month)

    for col in range(1, max_col_pc):
        if cell_pc(1, col).value == month_phkt[month_index]:
            selected_col = get_column_letter(col + 1)

    new.odoo_sheet.column_dimensions[get_column_letter(max_col_odoo + 1)].width = 5

    for col in range(4):
        column = max_col_odoo + col + 2
        for row in range(4, max_row_odoo + 1):
            if row == 4:
                cell_odoo(row, column).value = header[col]
            elif 4 < row < max_row_odoo:
                if col == 0:
                    cell_odoo(row, column).value = f"=sum(G{row}:K{row})"
                elif col == 1:
                    cell_odoo(row,
                              column).value = f"=IFERROR(VLOOKUP(B{row},CHOOSE(" + '{1,2},' + \
                                              f"'{new.wb.sheetnames[0]}'!$F$6:$F$800,'{new.wb.sheetnames[0]}'!" \
                                              f"${selected_col}$6:${selected_col}$800),2,0),0)"
                elif col == 2:
                    cell_odoo(row,
                              column).value = f'={cell_odoo(row, column - 2).coordinate}-' \
                                              f'{cell_odoo(row, column - 1).coordinate}'
            elif row == max_row_odoo:
                if col == 3:
                    pass
                else:
                    loc = get_column_letter(column)
                    cell_odoo(row, column).value = f"=sum({loc}5:{loc}{max_row_odoo - 1})"
            new.odoo_sheet.column_dimensions[get_column_letter(column)].width = 20
            cell_odoo(row, column).number_format = copy(cell_odoo(row, max_col_odoo).number_format)
            cell_odoo(row, column).font = copy(cell_odoo(row, max_col_odoo).font)
            cell_odoo(row, column).border = copy(cell_odoo(row, max_col_odoo).border)
            cell_odoo(row, column).fill = copy(cell_odoo(row, max_col_odoo).fill)
            cell_odoo(row, column).protection = copy(cell_odoo(row, max_col_odoo).protection)
            cell_odoo(row, column).alignment = copy(cell_odoo(row, max_col_odoo).alignment)
    max_col_odoo = new.odoo_sheet.max_column

    for col in range(8, 13):
        empty = True
        for row in range(5, max_row_odoo):
            if cell_odoo(row, col).value != 0:
                empty = False
                break
        if empty:
            new.odoo_sheet.column_dimensions[get_column_letter(col)].hidden = True

    for col in range(5, 13):
        cell_odoo(max_row_odoo,
                  col).value = f'=SUM({get_column_letter(col)}5:{get_column_letter(col)}{max_row_odoo - 1})'

    for col in range(5, max_col_odoo + 5):
        for row in range(5, max_row_odoo + 1):
            cell_odoo(row, col).number_format = '#,##0.00;[RED]- #,##0.00;-'


def _quit():
    quit()
    sys.exit()


def password_entry():
    pass_window = Tk()
    password_label = Label(pass_window, text='Password')
    password_label.pack()
    pass_window.mainloop()


window = Win()
header = ['Odoo', 'Excel', 'Difference', 'Remarks']
month_number = ['01', '02', '03', '04', '05', '06', '07', ' 08', '09', '10', '11', '12']
month_eni = ["JAN", "FEB", "MAR", "APR", "MEI", "JUNI", "JULI", "AGST", "SEPT", "OKT", "NOV", "DES"]
month_phkt = ['JANUARI 2021', 'FEBRUARI 2021', 'MARET 2021', 'APRIL 2021', 'MEI 2021', 'JUNI 2021',
              'JULI 2021', 'AGUSTUS 2021', 'SEPTEMBER 2021', 'OKTOBER 2021', 'NOVEMBER 2021', 'DESEMBER 2021']
month_odoo = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', "Sep", 'Oct', 'Nov', 'Dec']
selected_col = ''
project = ''

ws = window.winfo_screenwidth()
hs = window.winfo_screenheight()
x = (ws / 2) - 275
y = (hs / 2) - 225
window.geometry(f"550x450+{int(x)}+{int(y)}")
window.configure(bg="#ffffff")
canvas = Canvas(
    window,
    bg="#ffffff",
    height=450,
    width=550,
    bd=0,
    highlightthickness=0,
    relief="ridge")
canvas.place(x=0, y=0)

background_img = PhotoImage(file=f"win/img/background.png")
background = canvas.create_image(
    275.0, 225.0,
    image=background_img)

pc = Load(window, canvas)
odoo = Load(window, canvas)
new = Create()

pc_icon = PhotoImage(file=f"win/img/img0.png")
pc.button.config(image=pc_icon)
pc.button.place(
    x=74, y=219,
    width=284,
    height=28)
pc.label = canvas.create_text(500, 254, text='', font=('calibri light', 8), fill='white', anchor=E)
pc.options.place(x=352, y=219, width=155, height=28)

odoo_icon = PhotoImage(file=f"win/img/img1.png")
odoo.button.config(image=odoo_icon)
odoo.button.place(
    x=74, y=275,
    width=284,
    height=28)
odoo.label = canvas.create_text(500, 309, text='', font=('calibri light', 8), fill='white', anchor=E)
odoo.options.place(x=352, y=275, width=155, height=28)

help_icon = PhotoImage(file=f"win/img/img2.png")
help_button = Button(
    image=help_icon,
    borderwidth=0,
    highlightthickness=0,
    command=_help,
    relief="flat")
help_button.place(
    x=3, y=420,
    width=24,
    height=24)

start_image = PhotoImage(file=f"win/img/img3.png")
start_button = Button(
    image=start_image,
    borderwidth=0,
    highlightthickness=0,
    command=start,
    relief="flat",
    state=DISABLED)
start_button.place(
    x=418, y=349,
    width=88,
    height=28)

close_image = PhotoImage(file=f"win/img/img4.png")
close_button = Button(
    image=close_image,
    borderwidth=0,
    highlightthickness=0,
    command=_quit,
    relief="flat")
close_button.place(
    x=524, y=6,
    width=19,
    height=19)

window.resizable(False, False)
window.mainloop()
