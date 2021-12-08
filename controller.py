import sys
from excel.helper import *
from excel.var import *
from zipfile import BadZipfile
from tkinter.messagebox import showinfo, showwarning
from tkinter.filedialog import asksaveasfilename
from openpyxl import *
from datetime import *
from os import startfile
from excel.copy_sheet_data import *
from excel.adjust import general, eni, phkt, phm_edi
import pywintypes


def exit_app():
    quit()
    sys.exit()


def home_show(self):
    setting_hide(self)
    guide_hide(self)
    self.canvas.create_image(
        140, 92,
        anchor="nw",
        image=self.main_title_img,
        tags="home")

    self.canvas.create_image(
        140, 310,
        anchor="nw",
        image=self.label_canvas,
        tags="home")

    self.canvas.create_image(
        140, 390,
        anchor="nw",
        image=self.label_canvas,
        tags="home")

    self.canvas.create_rectangle(
        370, 350, 370 + 120, 350 + 40,
        fill="#81c0f2",
        outline="",
        tags="home")

    self.canvas.create_rectangle(
        370, 270, 370 + 120, 270 + 40,
        fill="#81c0f2",
        outline="",
        tags="home")

    self.canvas.create_text(
        405.0, 290.0,
        text="Pass: ",
        fill="#ffffff",
        font=("KarlaTamilUpright-Bold", int(11.0)),
        tags="home")

    pc_label = pc.label
    self.pc_label = self.canvas.create_text(
        150.0, 322.0,
        fill="#ffffff",
        anchor="w",
        font=("KarlaTamilUpright-Regular", int(9.0)),
        tags="home")

    odoo_label = odoo.label
    self.odoo_label = self.canvas.create_text(
        150.0, 401.0,
        fill="#ffffff",
        anchor="w",
        font=("KarlaTamilUpright-Regular", int(9.0)),
        tags="home")

    self.pc_button.place(
        x=140, y=270,
        width=230,
        height=40)

    self.pc_options.place(x=490, y=270, width=120, height=40)

    self.entry0.place(
        x=430.0, y=280,
        width=36.0,
        height=22)

    self.odoo_button.place(
        x=140, y=350,
        width=230,
        height=40)

    self.odoo_options.place(x=490, y=350, width=120, height=40)

    self.start_button.place(
        x=660, y=460,
        width=100,
        height=40)


def home_hide(self):
    self.canvas.delete("home")
    self.pc_button.place_forget()
    self.pc_options.place_forget()
    self.entry0.place_forget()
    self.odoo_button.place_forget()
    self.odoo_options.place_forget()
    self.start_button.place_forget()


def guide_show(self):
    home_hide(self)
    setting_hide(self)
    self.canvas.create_image(
        140, 96,
        anchor="nw",
        image=self.help_title_img,
        tags="guide")

    self.canvas.create_text(
        150.0, 210.0,
        text="1. Download form monitoring yang dikirimkan Project Controller\n"
             "2. Download sales report dari Odoo, pastikan tanggal yang dipilih\n"
             "    mencakup 1 bulan penuh\n"
             '3. Klik tombol "Monitoring by PC", pilih file yang diinginkan\n'
             '4. Klik tombol "Monitoring Odoo", pilih file yang diingingkan\n'
             "5. Pilih sheet monitoring yang diperlukan untuk masing-masing file\n"
             '6. Klik "Start"\n'
             "7. Simpan file output sesuai format yang diinginkan\n"
             "8. Adjust file output sesuai kebutuhan\n"
             "9. Selesai",
        fill="#ffffff",
        anchor="nw",
        font=("Cabin-Regular", int(14.0)),
        tags="guide")


def guide_hide(self):
    self.canvas.delete("guide")


def setting_show(self):
    home_hide(self)
    guide_hide(self)
    self.canvas.create_image(
        140, 96,
        anchor="nw",
        image=self.setting_title_img,
        tags="setting")

    self.canvas.create_image(
        140, 310,
        anchor="nw",
        image=self.label_canvas,
        tags="setting")

    self.canvas.create_image(
        140, 390,
        anchor="nw",
        image=self.label_canvas,
        tags="setting")

    self.open_d_button.place(
        x=140, y=270,
        width=230,
        height=40)

    self.save_d_button.place(
        x=140, y=350,
        width=230,
        height=40)


def setting_hide(self):
    self.canvas.delete('setting')
    self.open_d_button.place_forget()
    self.save_d_button.place_forget()


def open_excel(app, main_source, aux_source, options, label):
    try:
        main_source.load(app, options, label)
    except BadZipfile:
        try:
            main_source.path = unlock_excel(main_source.path, str(app.password.get()))
            main_source.workbook = load_workbook(main_source.path, data_only=True)
            main_source.load_attribute(app, options, label)
        #
        except pywintypes.com_error:
            showwarning(title='Incorrect Password',
                        message='Incorrect Password!\n'
                                'Please make sure you type in correct Password')

        # except AttributeError: ("Error")

    if main_source.state and aux_source.state:
        app.start_button.config(state="normal")


def start(app, new_excel, pc_excel, odoo_excel):
    new_excel.create_excel(app)

    pc_excel.active_sheet = pc_excel.workbook[app.pc_options.get()]
    copy_sheet(pc_excel.active_sheet, new_excel.pc_sheet)
    odoo_excel.active_sheet = odoo_excel.workbook[app.odoo_options.get()]
    copy_sheet(odoo_excel.active_sheet, new_excel.odoo_sheet)
    if 'Sheet' in new_excel.sheet_list:
        new_excel.workbook.remove(new_excel.workbook['Sheet'])
    if new_excel.workbook.sheetnames[1] == '3250':
        new_excel.month_pc = ["JAN", "FEB", "MAR", "APR", "MEI", "JUNI", "JULI", "AGST", "SEPT", "OKT", "NOV", "DES"]
        eni(new_excel)
        new.project = 'ENI'
    elif new_excel.workbook.sheetnames[1] == '3235':
        new_excel.month_pc = ["2021-01", "2021-02", "2021-03", "2021-04", "2021-05", "2021-06", "2021-07", "2021-08",
                              "2021-09", "2021-10", "2021-11", "2021-12"]
        phm_edi(new_excel)
        project = 'PHM'
    elif new_excel.workbook.sheetnames[1] == '3247':
        new_excel.month_pc = ['JANUARI 2021', 'FEBRUARI 2021', 'MARET 2021', 'APRIL 2021', 'MEI 2021',
                              'JUNI 2021', 'JULI 2021', 'AGUSTUS 2021', 'SEPTEMBER 2021', 'OKTOBER 2021',
                              'NOVEMBER 2021', 'DESEMBER 2021']
        phkt(new_excel)
        project = 'PHKT'

    dates = datetime.now()
    if int(dates.strftime("%d")) < 8:
        week = "01"
    elif 7 < int(dates.strftime("%d")) < 15:
        week = "02"
    elif 14 < int(dates.strftime("%d")) < 22:
        week = "03"
    elif 21 < int(dates.strftime("%d")):
        week = "04"

    path = asksaveasfilename(
        initialdir='C:/Users/mohin/Downloads/',
        title='Save File',
        initialfile=f'{new_excel.workbook.sheetnames[1]}-{dates.strftime("%Y%m%d")}-W{week}-Monitoring-v01.xlsx',
        filetypes=(("Excel File", "*xlsx"), ("All Files", "*.*")))
    if path != '':
        new_excel.workbook.save(path)
        showinfo(title='Done', message='Done!')
        startfile(path)
    else:
        return
