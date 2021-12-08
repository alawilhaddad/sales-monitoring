from win.config import Win
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import *
from excel.helper import convert_slash, remove_password_xlsx
from openpyxl import *
from zipfile import BadZipfile
from win.funct import *


def home_show():
    setting_hide()
    guide_hide()
    canvas.create_image(
        140, 92,
        anchor="nw",
        image=main_title_img,
        tags="home")

    canvas.create_image(
        140, 310,
        anchor="nw",
        image=label_canvas,
        tags="home")

    canvas.create_image(
        140, 390,
        anchor="nw",
        image=label_canvas,
        tags="home")

    canvas.create_rectangle(
        370, 350, 370 + 120, 350 + 40,
        fill="#81c0f2",
        outline="",
        tags="home")

    canvas.create_rectangle(
        370, 270, 370 + 120, 270 + 40,
        fill="#81c0f2",
        outline="",
        tags="home")

    canvas.create_text(
        405.0, 290.0,
        text="Pass: ",
        fill="#ffffff",
        font=("KarlaTamilUpright-Bold", int(11.0)),
        tags="home")

    pc.label = canvas.create_text(
        150.0, 322.0,
        fill="#ffffff",
        anchor="w",
        font=("KarlaTamilUpright-Regular", int(9.0)),
        tags="home")

    odoo.label = canvas.create_text(
        150.0, 401.0,
        fill="#ffffff",
        anchor="w",
        font=("KarlaTamilUpright-Regular", int(9.0)),
        tags="home")

    pc.button.place(
        x=140, y=270,
        width=230,
        height=40)

    pc.options.place(x=490, y=270, width=120, height=40)

    entry0.place(
        x=430.0, y=280,
        width=36.0,
        height=22)

    odoo.button.place(
        x=140, y=350,
        width=230,
        height=40)

    odoo.options.place(x=490, y=350, width=120, height=40)

    start_button.place(
        x=660, y=460,
        width=100,
        height=40)


def home_hide():
    canvas.delete("home")
    pc.button.place_forget()
    pc.options.place_forget()
    entry0.place_forget()
    odoo.button.place_forget()
    odoo.options.place_forget()
    start_button.place_forget()


def guide_show():
    home_hide()
    setting_hide()
    canvas.create_image(
        140, 96,
        anchor="nw",
        image=help_title_img,
        tags="guide")

    canvas.create_text(
        150.0, 210.0,
        text="1. Download form monitoring yang dikirimkan Project Controller\n"
             "2. Download sales report dari Odoo, pastikan tanggal yang dipilih\n"
             "    mencakup 1 bulan penuh\n"
             '3. Klik tombol "Monitoring by PC", pilih file yang diinginkan\n'
             '4. Klik tombol "Monitoring Odoo", pilih file yang diingingkan\n'
             "5. Pilih sheet monitoring yang diperlukan untuk masing-masing file\n"
             '6. Klik "Start"\n'
             "7. Simpan file output sesuai format yang diinginkan\n"
             "8. Adjust file output sesuai kebutuhan\n"
             "9. Selesai",
        fill="#ffffff",
        anchor="nw",
        font=("Cabin-Regular", int(14.0)),
        tags="guide")


def guide_hide():
    canvas.delete("guide")


def setting_show():
    home_hide()
    guide_hide()
    canvas.create_image(
        140, 96,
        anchor="nw",
        image=setting_title_img,
        tags="setting")

    canvas.create_image(
        140, 310,
        anchor="nw",
        image=label_canvas,
        tags="setting")

    canvas.create_image(
        140, 390,
        anchor="nw",
        image=label_canvas,
        tags="setting")

    open_d_button.place(
        x=140, y=270,
        width=230,
        height=40)

    save_d_button.place(
        x=140, y=350,
        width=230,
        height=40)


def setting_hide():
    canvas.delete('setting')
    open_d_button.place_forget()
    save_d_button.place_forget()


def _open(self, cvs):
    self.path = askopenfilename(initialdir='C:/Users/mohin/Downloads/',
                                title='Open File',
                                filetypes=(("Excel File", "*xlsx"), ("All Files", "*.*")))
    self.filename = self.path.split('/')[-1]
    self.path = convert_slash(self.path)
    if self.path is None:
        return
    else:
        try:
            self.wb = load_workbook(self.path, data_only=True)
            self.sheet_list = self.wb.sheetnames
            self.active_sheet = self.sheet_list[0]
            cvs.itemconfigure(self.label, text=self.filename)
            self.options['values'] = self.sheet_list
            self.options['state'] = 'readonly'
            self.sheet.set(self.sheet_list[0])
            self.state = True
        except BadZipfile:
            showinfo(title='Password Protected', message='Your file is password protected.')
            self.path = remove_password_xlsx(self.path, input("password: "))
            self.filename = self.path.split('\\')[-1]
            self.wb = load_workbook(self.path, data_only=True)
            self.sheet_list = self.wb.sheetnames
            self.active_sheet = self.sheet_list[0]
            canvas.itemconfigure(self.label, text=self.filename)
            self.options['values'] = self.sheet_list
            self.options['state'] = 'readonly'
            self.sheet.set(self.sheet_list[0])
            self.state = True
        if pc.state and odoo.state:
            start_button.config(state=NORMAL)


if __name__ == "__main__":
    window = Win()
    ws = window.winfo_screenwidth()
    hs = window.winfo_screenheight()
    x = (ws / 2) - (window_width / 2)
    y = (hs / 2) - (window_height / 2)
    window.geometry(f"{window_width}x{window_height}+{int(x)}+{int(y)}")
    canvas = Canvas(
        window,
        bg="#ffffff",
        height=600,
        width=800,
        bd=0,
        highlightthickness=0,
        relief="ridge")
    canvas.place(x=0, y=0)

    background_img = PhotoImage(file=f"win/img/background.png")
    background = canvas.create_image(400, 300, image=background_img)

    home_icon = PhotoImage(file=f"win/img/home_icon.png")
    home_button = Button(
        image=home_icon,
        borderwidth=0,
        highlightthickness=0,
        command=home_show,
        relief="flat")
    home_button.place(
        x=30, y=370,
        width=60,
        height=60)

    setting_icon = PhotoImage(file=f"win/img/setting_icon.png")
    setting_button = Button(
        image=setting_icon,
        borderwidth=0,
        highlightthickness=0,
        command=setting_show,
        relief="flat")
    setting_button.place(
        x=30, y=440,
        width=60,
        height=60)

    help_icon = PhotoImage(file=f"win/img/help_icon.png")
    help_button = Button(
        image=help_icon,
        borderwidth=0,
        highlightthickness=0,
        command=guide_show,
        relief="flat")
    help_button.place(
        x=30, y=510,
        width=60,
        height=60)

    canvas.create_text(
        155.0, 562.0,
        text="Haddaddegusti 2021 | v. 2.3.0",
        fill="#ffffff",
        anchor="w",
        font=("Taprom", int(10.0)))

    close_image = PhotoImage(file=f"win/img/close_icon.png")
    close_button = Button(
        image=close_image,
        borderwidth=0,
        highlightthickness=0,
        command=exit_app,
        relief="flat")
    close_button.place(
        x=750, y=30,
        width=20,
        height=20)

    # ===================================================Home====================================================

    main_title_img = PhotoImage(file=f"win/img/main_title.png")

    label_canvas = PhotoImage(file=f"win/img/label_canvas.png")

    pc.image = PhotoImage(file=f"win/img/pc_icon.png")
    pc.button = Button(
        image=pc.image,
        borderwidth=0,
        highlightthickness=0,
        relief="flat")
    pc.button['command'] = lambda source=pc, cv=canvas: _open(source, cv)

    pc.sheet = StringVar(window)
    pc.options = ttk.Combobox(
        window,
        textvariable=pc.sheet,
        width=12,
        value=pc.ws,
        font=('karla tamil upright', 9),
        justify='right',
        state="readonly")

    pc.options['state'] = 'disabled'

    entry0_img = PhotoImage(file=f"win/img/img_textBox0.png")
    entry0 = Entry(
        bd=0,
        fg="#7e91d6",
        bg="#ffffff",
        highlightthickness=0,
        font=("KarlaTamilUpright-Regular", int(11.0)))

    odoo.image = PhotoImage(file=f"win/img/odoo_icon.png")
    odoo.button = Button(
        image=odoo.image,
        borderwidth=0,
        highlightthickness=0,
        relief="flat")
    odoo.button['command'] = lambda source=odoo, cv=canvas: _open(source, cv)

    odoo.sheet = StringVar(window)
    odoo.options = ttk.Combobox(
        window,
        textvariable=odoo.sheet,
        width=12,
        value=odoo.ws,
        font=('karla tamil upright', 9),
        justify='right',
        state="readonly")

    odoo.options['state'] = 'disabled'

    start_image = PhotoImage(file=f"win/img/start_icon.png")
    start_button = Button(
        image=start_image,
        borderwidth=0,
        highlightthickness=0,
        command=start,
        relief="flat",
        state=DISABLED)

    help_title_img = PhotoImage(file=f"win/img/help_title.png")

    setting_title_img = PhotoImage(file=f"win/img/setting_title.png")

    open_d_image = PhotoImage(file=f"win/img/open_directory.png")
    open_d_button = Button(
        image=open_d_image,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: print("click"),
        relief="flat")

    save_d_image = PhotoImage(file=f"win/img/save_directory.png")
    save_d_button = Button(
        image=save_d_image,
        borderwidth=0,
        highlightthickness=0,
        command=lambda: print("click"),
        relief="flat")

    home_show()
    window.mainloop()
